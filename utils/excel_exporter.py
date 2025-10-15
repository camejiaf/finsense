"""
Excel export functionality for FinSense DCF and sentiment analysis
Creates comprehensive Excel workbooks with multiple sheets
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, List
import io


class ExcelExporter:
    """Export DCF and sentiment analysis results to Excel"""

    def __init__(self):
        self.workbook = None
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_comprehensive_analysis(self,
                                      stock_data: Dict,
                                      dcf_results: Dict,
                                      sentiment_summary: Dict,
                                      analyzed_articles: List[Dict],
                                      output_path: str = None) -> str:
        """
        Export comprehensive analysis to Excel with multiple sheets

        Args:
            stock_data: Stock information and financial data
            dcf_results: DCF calculation results
            sentiment_summary: Sentiment analysis summary
            analyzed_articles: List of analyzed news articles
            output_path: Optional custom output path

        Returns:
            Path to generated Excel file
        """

        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ticker = stock_data.get('ticker', 'UNKNOWN')
            output_path = f"FinSense_Analysis_{ticker}_{timestamp}.xlsx"

        # Create workbook
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

        # Create sheets
        self._create_summary_sheet(stock_data, dcf_results, sentiment_summary)
        self._create_dcf_details_sheet(dcf_results)
        self._create_sentiment_sheet(sentiment_summary, analyzed_articles)
        self._create_financials_sheet(stock_data)
        self._create_monte_carlo_sheet(dcf_results)

        # Save workbook
        self.workbook.save(output_path)

        return output_path

    def _create_summary_sheet(self, stock_data: Dict, dcf_results: Dict, sentiment_summary: Dict):
        """Create executive summary sheet"""

        ws = self.workbook.create_sheet("Executive Summary")

        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = f"FinSense Analysis - {stock_data.get('ticker', 'N/A')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].fill = self.header_fill
        ws['A1'].alignment = Alignment(horizontal='center')

        # Company Info
        ws['A3'] = "Company Information"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill

        company_data = [
            ["Ticker", stock_data.get('ticker', 'N/A')],
            ["Company Name", stock_data.get('company_name', 'N/A')],
            ["Current Price", f"${stock_data.get('current_price', 0):.2f}"],
            ["Market Cap", f"${stock_data.get('market_cap', 0)/1e9:.2f}B"],
            ["Shares Outstanding",
                f"{stock_data.get('shares_outstanding', 0)/1e9:.2f}B"]
        ]

        for i, (label, value) in enumerate(company_data, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)

        # DCF Summary
        ws['A10'] = "DCF Analysis Summary"
        ws['A10'].font = self.header_font
        ws['A10'].fill = self.header_fill

        if dcf_results and 'base_case_valuation' in dcf_results:
            base_case = dcf_results['base_case_valuation']
            monte_carlo = dcf_results.get('monte_carlo_results', {})

            dcf_data = [
                ["Intrinsic Value per Share",
                    f"${base_case.get('equity_value_per_share', 0):.2f}"],
                ["Enterprise Value",
                    f"${base_case.get('enterprise_value', 0)/1e9:.2f}B"],
                ["Monte Carlo Mean",
                    f"${monte_carlo.get('mean_valuation', 0):.2f}"],
                ["Monte Carlo Std Dev",
                    f"${monte_carlo.get('std_valuation', 0):.2f}"],
                ["5th Percentile",
                    f"${monte_carlo.get('percentile_5', 0):.2f}"],
                ["95th Percentile",
                    f"${monte_carlo.get('percentile_95', 0):.2f}"]
            ]

            for i, (label, value) in enumerate(dcf_data, 11):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)

            # Upside/Downside calculation
            current_price = stock_data.get('current_price', 0)
            dcf_price = base_case.get('equity_value_per_share', 0)

            if current_price > 0 and dcf_price > 0:
                upside = ((dcf_price - current_price) / current_price) * 100
                recommendation = "BUY" if upside > 10 else "HOLD" if upside > -10 else "SELL"

                ws['A18'] = "Recommendation"
                ws['B18'] = recommendation
                ws['A18'].font = Font(bold=True)
                ws['B18'].font = Font(bold=True, color="00AA00" if recommendation ==
                                      "BUY" else "FF6600" if recommendation == "HOLD" else "FF0000")

                ws['A19'] = "Upside/Downside"
                ws['B19'] = f"{upside:.1f}%"
                ws['A19'].font = Font(bold=True)

        # Sentiment Summary
        ws['A22'] = "News Sentiment Summary"
        ws['A22'].font = self.header_font
        ws['A22'].fill = self.header_fill

        if sentiment_summary and sentiment_summary.get('total_articles', 0) > 0:
            sentiment_data = [
                ["Total Articles", sentiment_summary.get('total_articles', 0)],
                ["Positive",
                    f"{sentiment_summary.get('positive_count', 0)} ({sentiment_summary.get('positive_percentage', 0):.1f}%)"],
                ["Negative",
                    f"{sentiment_summary.get('negative_count', 0)} ({sentiment_summary.get('negative_percentage', 0):.1f}%)"],
                ["Neutral",
                    f"{sentiment_summary.get('neutral_count', 0)} ({sentiment_summary.get('neutral_percentage', 0):.1f}%)"],
                ["Overall Sentiment", sentiment_summary.get(
                    'overall_sentiment', 'neutral').title()],
                ["Average Confidence",
                    f"{sentiment_summary.get('average_confidence', 0):.1%}"]
            ]

            for i, (label, value) in enumerate(sentiment_data, 23):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Add borders
        self._add_borders_to_range(ws, 'A3:B30')

    def _create_dcf_details_sheet(self, dcf_results: Dict):
        """Create detailed DCF analysis sheet"""

        ws = self.workbook.create_sheet("DCF Details")

        # Header
        ws['A1'] = "DCF Valuation Details"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill

        if not dcf_results or 'base_case_valuation' not in dcf_results:
            ws['A3'] = "No DCF data available"
            return

        base_case = dcf_results['base_case_valuation']
        assumptions = dcf_results.get('assumptions', {})

        # Assumptions
        ws['A3'] = "Key Assumptions"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill

        assumption_data = [
            ["Base FCF", f"${assumptions.get('base_fcf', 0)/1e9:.2f}B"],
            ["FCF Growth Rate",
                f"{assumptions.get('fcf_growth_rate', 0)*100:.1f}%"],
            ["WACC", f"{assumptions.get('wacc', 0)*100:.1f}%"],
            ["Terminal Growth Rate",
                f"{assumptions.get('terminal_growth', 0)*100:.1f}%"],
            ["Shares Outstanding",
                f"{assumptions.get('shares_outstanding', 0)/1e9:.2f}B"]
        ]

        for i, (label, value) in enumerate(assumption_data, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)

        # FCF Projections
        ws['A10'] = "5-Year FCF Projections"
        ws['A10'].font = self.header_font
        ws['A10'].fill = self.header_fill

        ws['A11'] = "Year"
        ws['B11'] = "Projected FCF"
        ws['C11'] = "Present Value"
        ws['A11'].font = Font(bold=True)
        ws['B11'].font = Font(bold=True)
        ws['C11'].font = Font(bold=True)

        fcf_projections = base_case.get('fcf_projections', [])
        pv_projections = base_case.get('pv_fcf_projections', [])

        for i, (year, fcf, pv) in enumerate(zip(range(1, 6), fcf_projections, pv_projections), 12):
            ws[f'A{i}'] = year
            ws[f'B{i}'] = f"${fcf/1e9:.2f}B"
            ws[f'C{i}'] = f"${pv/1e9:.2f}B"

        # Terminal Value
        ws['A18'] = "Terminal Value Calculation"
        ws['A18'].font = self.header_font
        ws['A18'].fill = self.header_fill

        terminal_fcf = base_case.get('terminal_fcf', 0)
        terminal_value = base_case.get('terminal_value', 0)
        pv_terminal = base_case.get('pv_terminal_value', 0)

        terminal_data = [
            ["Terminal Year FCF", f"${terminal_fcf/1e9:.2f}B"],
            ["Terminal Value", f"${terminal_value/1e9:.2f}B"],
            ["PV of Terminal Value", f"${pv_terminal/1e9:.2f}B"]
        ]

        for i, (label, value) in enumerate(terminal_data, 19):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)

        # Summary
        ws['A23'] = "Valuation Summary"
        ws['A23'].font = self.header_font
        ws['A23'].fill = self.header_fill

        pv_explicit = base_case.get('pv_explicit_period', 0)
        enterprise_value = base_case.get('enterprise_value', 0)
        equity_value = base_case.get('equity_value_per_share', 0)

        summary_data = [
            ["PV of Explicit Period", f"${pv_explicit/1e9:.2f}B"],
            ["PV of Terminal Value", f"${pv_terminal/1e9:.2f}B"],
            ["Enterprise Value", f"${enterprise_value/1e9:.2f}B"],
            ["Equity Value per Share", f"${equity_value:.2f}"]
        ]

        for i, (label, value) in enumerate(summary_data, 24):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Add borders
        self._add_borders_to_range(ws, 'A3:B28')
        self._add_borders_to_range(ws, 'A10:C17')

    def _create_sentiment_sheet(self, sentiment_summary: Dict, analyzed_articles: List[Dict]):
        """Create sentiment analysis sheet"""

        ws = self.workbook.create_sheet("Sentiment Analysis")

        # Header
        ws['A1'] = "News Sentiment Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill

        # Summary
        ws['A3'] = "Sentiment Summary"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill

        if sentiment_summary and sentiment_summary.get('total_articles', 0) > 0:
            summary_data = [
                ["Total Articles", sentiment_summary.get('total_articles', 0)],
                ["Positive Count", sentiment_summary.get('positive_count', 0)],
                ["Negative Count", sentiment_summary.get('negative_count', 0)],
                ["Neutral Count", sentiment_summary.get('neutral_count', 0)],
                ["Positive %",
                    f"{sentiment_summary.get('positive_percentage', 0):.1f}%"],
                ["Negative %",
                    f"{sentiment_summary.get('negative_percentage', 0):.1f}%"],
                ["Neutral %",
                    f"{sentiment_summary.get('neutral_percentage', 0):.1f}%"],
                ["Overall Sentiment", sentiment_summary.get(
                    'overall_sentiment', 'neutral').title()],
                ["Average Confidence",
                    f"{sentiment_summary.get('average_confidence', 0):.1%}"]
            ]

            for i, (label, value) in enumerate(summary_data, 4):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)

        # Detailed Articles
        if analyzed_articles:
            ws['A14'] = "Article Details"
            ws['A14'].font = self.header_font
            ws['A14'].fill = self.header_fill

            # Headers
            headers = ["Title", "Sentiment", "Confidence",
                       "Positive Score", "Negative Score", "Published"]
            for i, header in enumerate(headers, 1):
                ws.cell(row=15, column=i, value=header)
                ws.cell(row=15, column=i).font = Font(bold=True)
                ws.cell(row=15, column=i).fill = self.header_fill

            # Article data
            # Limit to 20 articles
            for i, article in enumerate(analyzed_articles[:20], 16):
                ws.cell(row=i, column=1, value=article.get(
                    'title', '')[:100])  # Truncate title
                ws.cell(row=i, column=2, value=article.get('sentiment', ''))
                ws.cell(row=i, column=3,
                        value=f"{article.get('confidence', 0):.2f}")
                ws.cell(row=i, column=4,
                        value=f"{article.get('positive_score', 0):.2f}")
                ws.cell(row=i, column=5,
                        value=f"{article.get('negative_score', 0):.2f}")
                ws.cell(row=i, column=6, value=article.get(
                    'published', '')[:20])

            # Format columns
            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20

            # Add borders
            self._add_borders_to_range(ws, 'A15:F35')

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Add borders
        self._add_borders_to_range(ws, 'A3:B12')

    def _create_financials_sheet(self, stock_data: Dict):
        """Create financial data sheet"""

        ws = self.workbook.create_sheet("Financial Data")

        # Header
        ws['A1'] = "Financial Data"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill

        # FCF Data
        ws['A3'] = "Free Cash Flow History"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill

        fcf_data = stock_data.get('fcf_data', [])
        if fcf_data:
            ws['A4'] = "Year"
            ws['B4'] = "Free Cash Flow"
            ws['A4'].font = Font(bold=True)
            ws['B4'].font = Font(bold=True)

            for i, fcf in enumerate(fcf_data, 5):
                ws[f'A{i}'] = f"Year {i}"
                ws[f'B{i}'] = f"${fcf/1e9:.2f}B"

        # Growth Rate
        fcf_growth = stock_data.get('fcf_growth_rate', 0)
        if fcf_growth != 0:
            ws['A10'] = "FCF Growth Rate"
            ws['B10'] = f"{fcf_growth*100:.1f}%"
            ws['A10'].font = Font(bold=True)

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Add borders
        self._add_borders_to_range(ws, 'A3:B15')

    def _create_monte_carlo_sheet(self, dcf_results: Dict):
        """Create Monte Carlo simulation results sheet"""

        ws = self.workbook.create_sheet("Monte Carlo")

        # Header
        ws['A1'] = "Monte Carlo Simulation Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill

        if not dcf_results or 'monte_carlo_results' not in dcf_results:
            ws['A3'] = "No Monte Carlo data available"
            return

        monte_carlo = dcf_results['monte_carlo_results']

        # Statistics
        ws['A3'] = "Monte Carlo Statistics"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill

        stats_data = [
            ["Mean Valuation", f"${monte_carlo.get('mean_valuation', 0):.2f}"],
            ["Median Valuation",
                f"${monte_carlo.get('median_valuation', 0):.2f}"],
            ["Standard Deviation",
                f"${monte_carlo.get('std_valuation', 0):.2f}"],
            ["5th Percentile", f"${monte_carlo.get('percentile_5', 0):.2f}"],
            ["25th Percentile", f"${monte_carlo.get('percentile_25', 0):.2f}"],
            ["75th Percentile", f"${monte_carlo.get('percentile_75', 0):.2f}"],
            ["95th Percentile", f"${monte_carlo.get('percentile_95', 0):.2f}"]
        ]

        for i, (label, value) in enumerate(stats_data, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)

        # All valuations (for charting)
        all_valuations = monte_carlo.get('all_valuations', [])
        if all_valuations:
            ws['A12'] = "All Valuations (Sample)"
            ws['A12'].font = self.header_font
            ws['A12'].fill = self.header_fill

            ws['A13'] = "Run"
            ws['B13'] = "Valuation"
            ws['A13'].font = Font(bold=True)
            ws['B13'].font = Font(bold=True)

            # Show first 100 valuations
            for i, valuation in enumerate(all_valuations[:100], 14):
                ws[f'A{i}'] = i
                ws[f'B{i}'] = f"${valuation:.2f}"

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Add borders
        self._add_borders_to_range(ws, 'A3:B10')
        if all_valuations:
            self._add_borders_to_range(ws, 'A12:B113')

    def _add_borders_to_range(self, ws, range_str):
        """Add borders to a range of cells"""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).border = self.border

    def export_simple_analysis(self,
                               ticker: str,
                               current_price: float,
                               dcf_price: float,
                               sentiment_summary: Dict,
                               output_path: str = None) -> str:
        """
        Export simple analysis to Excel

        Args:
            ticker: Stock ticker symbol
            current_price: Current stock price
            dcf_price: DCF calculated price
            sentiment_summary: Sentiment analysis summary
            output_path: Optional output path

        Returns:
            Path to generated Excel file
        """

        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"FinSense_QuickAnalysis_{ticker}_{timestamp}.xlsx"

        # Create simple DataFrame
        data = {
            'Metric': [
                'Ticker',
                'Current Price',
                'DCF Value',
                'Recommendation',
                'Upside/Downside'
            ],
            'Value': [
                ticker,
                f"${current_price:.2f}",
                f"${dcf_price:.2f}",
                "BUY" if dcf_price > current_price *
                1.1 else "HOLD" if dcf_price > current_price * 0.9 else "SELL",
                f"{((dcf_price - current_price) / current_price * 100):.1f}%" if current_price > 0 else "N/A"
            ]
        }

        if sentiment_summary and sentiment_summary.get('total_articles', 0) > 0:
            data['Metric'].extend([
                'News Sentiment',
                'Positive %',
                'Articles Analyzed'
            ])
            data['Value'].extend([
                sentiment_summary.get('overall_sentiment', 'neutral').title(),
                f"{sentiment_summary.get('positive_percentage', 0):.1f}%",
                sentiment_summary.get('total_articles', 0)
            ])

        df = pd.DataFrame(data)

        # Export to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Analysis', index=False)

            # Format the sheet
            wb = writer.book
            ws = writer.sheets['Analysis']

            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = self.header_fill

            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15

        return output_path
